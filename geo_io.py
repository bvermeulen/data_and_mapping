import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date, timedelta
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
from shapely.ops import cascaded_union
from Utils.plogger import Logger
import inspect
import re


PREFIX = r'autoseis_data\OUT_'

# start logger
logformat = '%(asctime)s - %(levelname)s - %(message)s'
Logger.set_logger('autoseis.log', logformat, 'INFO')
logger = Logger.getlogger()

class GeoData:
    '''  method for handling Geo data '''
    def __init__(self):
        pass

    def read_geo_data(self, _date):
        read_is_valid = False
        _geo_file = ''.join([PREFIX, 
                            f'{_date.year:04}', f'{_date.month:02}', f'{_date.day:02}', 
                            '*.xlsx'])
        _geo_file = glob.glob(_geo_file)
        logger.info(f'filename: {_geo_file}')

        if len(_geo_file) == 1:
            try:
                self.geo_df = pd.read_excel(_geo_file[0])
                self.date = _date
                self.add_bat_days_in_field_to_df()
                read_is_valid = True

            except FileNotFoundError:
                logger.info(f'{inspect.stack()[0][3]} - Exception FileNotFoundError": {_geo_file[0]}')
        else:
            pass
        
        if read_is_valid:
            return True, self.geo_df
        else:    
            return False, None

    def add_bat_days_in_field_to_df(self):
        days_in_field = []
        for _, row in self.geo_df.iterrows():
            bat_start = str(row['BATSTART'])
            try:
                _year = int(bat_start[0:4])
                _julianday = int(bat_start[4:7])
                _date_bat_start = date(_year, 1, 1) + timedelta(_julianday)
            except ValueError:
                logger.info(f'{inspect.stack()[0][3]} - Exception ValueError - bat_start: '
                            f'{bat_start}')
                _date_bat_start = date(1900, 1, 1)
            
            bat_start_new = str(row['BATSTART_NEW'])
            try:
                _year = int(bat_start_new[0:4])
                _julianday = int(bat_start_new[4:7])
                _date_bat_start_new = date(_year, 1, 1) + timedelta(_julianday)
            except ValueError:
                _date_bat_start_new = date(1900, 1, 1)

            _date_bat_start = max(_date_bat_start_new, _date_bat_start)
            if _date_bat_start != date(1900, 1, 1):
                _days_in_field = (self.date - _date_bat_start).days
            else:
                _days_in_field = np.NaN
            
            days_in_field.append(_days_in_field)

        # add the columns to the dataframe
        self.geo_df['days_in_field'] = days_in_field


def get_date():
    _date = input('date (YYMMDD) [q - quit]: ')
    if _date in ['q', 'Q']:
        exit()
    _date = date(int(_date[0:2])+2000, 
                 int(_date[2:4]), 
                 int(_date[4:6]))

    return _date


def daterange(start_date, end_date):
    for n in range(int ((end_date - start_date).days)):
        yield start_date + timedelta(n)


def get_date_range():
    start_date = input('date (YYMMDD) [q - quit]: ')
    end_date = input('date (YYMMDD) [q - quit]: ')
    if start_date in ['q', 'Q'] or end_date in ['q', 'Q']:
        exit()
    start_date = date(int(start_date[0:2])+2000, 
                      int(start_date[2:4]), 
                      int(start_date[4:6]))
    end_date = date(int(end_date[0:2])+2000, 
                      int(end_date[2:4]), 
                      int(end_date[4:6]))
    end_date += timedelta(1)

    if start_date >= end_date:
        print('incorrect date range')
        start_date = -1
        end_date = -1

    return start_date, end_date


def append_df_to_excel(df, filename, sheet_name='Sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
    :filename: File path or existing ExcelWriter
               (Example: '/path/to/file.xlsx')
    :df: dataframe to save to workbook
    :sheet_name: Name of sheet which will contain DataFrame.
                 (default: 'Sheet1')
    :startrow: upper left cell row to dump data frame.
               Per default (startrow=None) calculate the last row
               in the existing DF and write to the next row...
    :to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                      [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

            # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        logger.info(f'{inspect.stack()[0][3]} - Exception FileNotFoundError {filename}')

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def transformation(point):
    ''' transformation from RL-RP to Easting/ Northting
        Origin point taken from SPS receiver point 3400-4213
        Azimuth angle of prospect is 30 dregrees

        :input: point as tuple (RL, RP)
        :output: point as tuple (Easting, Northing)
    '''
    azimuth = (np.pi * 30 / 180 )  # converted to radians
    dx_crossline = 10.0 * np.cos(azimuth)
    dy_crossline = -10.0 * np.sin(azimuth)
    dx_inline = -10.0 * np.sin(azimuth)
    dy_inline = -10.0 * np.cos(azimuth)
    POINT_0 = (3400., 4213.)
    COORD_0 = (491074., 5358167.)

    transformed_point =  ((point[0] - POINT_0[0]) * dx_crossline +
                          (point[1] - POINT_0[1]) * dx_inline + COORD_0[0],
                          (point[0] - POINT_0[0]) * dy_crossline +
                          (point[1] - POINT_0[1]) * dy_inline + COORD_0[1])
    return transformed_point


def swath_selection():
    ''' Selection of swath. Swath information taken from the file: 
           Points+Lines_SW_24_stay.xlsx
        Manual inpput check if there exists at least one valid swath otherwise
        repeat the input

        retrieves the extent for each swath, transforms it to Easting,Northing and
        unions the swaths to one polynom

        :input: None
        :output: swaths_polynom of type Polynom of Shapely

    '''    
    swath_file = r'.\Points+Lines_SW_24_stay.xlsx'
    swath_df = pd.read_excel(swath_file, skiprows=5)
    valid_swaths = swath_df['Swath'].tolist()
    valid = False
    swaths = []
    while not valid:
        _swaths = [int(num[0]) for num in re.finditer(r'\d+', input('Swaths to be included: [0 for none]: '))]
        if len(_swaths) == 1 and _swaths[0] == 0:
            valid = True
            break
             
        for swath in _swaths:
            if swath in valid_swaths:
                swaths.append(swath)
        
        if swaths:
            valid = True

    swath_polygons = []
    for swath in swaths:
        sd = swath_df[swath_df['Swath'] == swath].iloc[0]

        point1 = (sd['1st RL'], sd['1st GP'])
        point2 = (sd['1st RL'], sd['last GP'])
        point3 = (sd['last RL'], sd['last GP'])
        point4 = (sd['last RL'], sd['1st GP'])
    
        point1_coord = transformation(point1)
        point2_coord = transformation(point2)
        point3_coord = transformation(point3)
        point4_coord = transformation(point4)
        
        swath_polygon = Polygon([point1_coord, point2_coord, point3_coord, point4_coord])
        swath_polygons.append(swath_polygon)

    return cascaded_union(swath_polygons)
